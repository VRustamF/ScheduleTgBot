# import pandas as pd
#
#
# df = pd.read_excel('test.xls', sheet_name='ФЭЭиУ-24', skiprows=7)
# df.to_excel('ФЭЭиУ-24.xlsx', index=False)

import xlrd
import openpyxl

# читаем xls
book = xlrd.open_workbook("test.xls", formatting_info=True)
sheet = book.sheet_by_name("ФЭЭиУ-24")

# создаём xlsx
wb = openpyxl.Workbook()
ws = wb.active

# копируем значения (не обязательно, но обычно нужно)
for r in range(7, sheet.nrows):
    for c in range(sheet.ncols):
        ws.cell(row=r-7+1, column=c+1).value = sheet.cell_value(r, c)

# переносим merged cells
for r1, r2, c1, c2 in sheet.merged_cells:
    # полностью выше данных → пропускаем
    if r2 <= 7:
        continue

    # подрезаем merge, если он начинается выше 8 строки
    new_r1 = max(r1, 7)
    new_r2 = r2

    ws.merge_cells(
        start_row=new_r1 - 7 + 1,
        end_row=new_r2 - 7,
        start_column=c1 + 1,
        end_column=c2
    )

wb.save("output.xlsx")