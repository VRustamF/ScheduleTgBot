import re
from collections import defaultdict
from pprint import pprint

from pydantic import BaseModel
import openpyxl as op
from openpyxl.cell.cell import MergedCell

filename = "output.xlsx"

########################################################################################################################

class Subjects(BaseModel):
    subject_name: str
    audience: str | None
    time: str

class DailySchedules(BaseModel):
    day_name: str
    daily_schedule: list[Subjects] = []

class WeeklySchedules(BaseModel):
    group_name: str
    schedule: list[DailySchedules] = []

class AllSchedules(BaseModel):
    schedules: list[WeeklySchedules] = []


BASE_SCHEDULE = AllSchedules()

########################################################################################################################

def add_subject(
        subject: Subjects,
        group_name: str,
        day_name: str,
):
    for weekly_schedule in BASE_SCHEDULE.schedules:
        if group_name == weekly_schedule.group_name:
            for day in weekly_schedule.schedule:
                if day_name == day.day_name:
                    day.daily_schedule.append(subject)
                    return


def add_day(
        day_name: str,
        group_name: str,
):
    for weekly_schedule in BASE_SCHEDULE.schedules:
        if group_name == weekly_schedule.group_name:
            weekly_schedule.schedule.append(DailySchedules(day_name=day_name))
            return

def add_schedule(
        schedule: WeeklySchedules,
):
    BASE_SCHEDULE.schedules.append(schedule)
    return

########################################################################################################################

wb = op.load_workbook(filename, data_only=True)
sheet = wb.active
sheet_columns = sheet.max_column
sheet_rows = sheet.max_row

day_column = ''
time_column = ''
groups_columns = defaultdict(str)

# Выдает значение ячейки даже если она объединенная
def get_merged_value(ws, row, col):
    cell = ws.cell(row=row, column=col)

    # обычная ячейка
    if not isinstance(cell, MergedCell):
        return cell.value

    # ищем merge-диапазон, в который она входит
    for cr in ws.merged_cells.ranges:
        if (cr.min_row <= row <= cr.max_row and
            cr.min_col <= col <= cr.max_col):
            return ws.cell(cr.min_row, cr.min_col).value

    return None

# Цикл, который пробегает по первой строке и ищет группы, колонку дней и колонку времени
for col in range(1, sheet_columns):
    cell_val = sheet.cell(row=1, column=col).value

    if cell_val[1:].isupper():
        group = cell_val
        new_weekly_schedule = WeeklySchedules(group_name=group)
        add_schedule(schedule=new_weekly_schedule)
        groups_columns[col] = group
        continue
    elif cell_val.lower() == 'время':
        time_column = col
    elif cell_val.lower().startswith('день'):
        day_column = col


# Парсит данные (вообще не оптимизировано, но мне пох)
current_day = ''
current_time = ''
added_day_in_groups = defaultdict(list)
for row in range(3, sheet_rows):
    for col in range(1, sheet_columns):
        cell_val = get_merged_value(ws=sheet, row=row, col=col)

        if cell_val and len(cell_val) > 2:
            if col == day_column and cell_val != current_day:
                current_day = cell_val
            elif col == time_column and cell_val != current_time:
                current_time = cell_val

            elif col in groups_columns.keys() and cell_val != get_merged_value(ws=sheet, row=row-1, col=col):

                if current_day not in added_day_in_groups[groups_columns[col]]:
                    add_day(day_name=current_day, group_name=groups_columns[col])
                    added_day_in_groups[groups_columns[col]].append(current_day)

                cleaned_val = re.sub(r'\s+', ' ', cell_val).strip()

                add_subject(
                    subject=Subjects(
                        subject_name=cleaned_val,
                        audience=sheet.cell(row=row, column=col+1).value,
                        time=current_time),
                    group_name=groups_columns[col],
                    day_name=current_day,
                )

pprint(BASE_SCHEDULE.model_dump())
