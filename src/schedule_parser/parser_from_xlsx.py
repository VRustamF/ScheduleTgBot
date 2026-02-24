import re
import asyncio
import logging
from pathlib import Path
from collections import defaultdict

import aiofiles
from aiofiles.os import listdir
import openpyxl as op
from openpyxl.cell.cell import MergedCell

from core import settings, BASE_DIR, db_helper
from crud.schedules import ScheduleService

logger = logging.getLogger(__name__)

pattern = re.compile(
    r"^([01]?\d|2[0-3]):[0-5]\d:[0-5]\d - ([01]?\d|2[0-3]):[0-5]\d:[0-5]\d$"
)


def get_merged_value(ws, r: int, c: int) -> str | None:
    """Выдает значение ячейки даже если она объединенная"""

    cell = ws.cell(row=r, column=c)

    # обычная ячейка
    if not isinstance(cell, MergedCell):
        return cell.value

    # ищем merge-диапазон, в который она входит
    for cr in ws.merged_cells.ranges:
        if cr.min_row <= r <= cr.max_row and cr.min_col <= c <= cr.max_col:
            return ws.cell(cr.min_row, cr.min_col).value

    return None


async def parser(form_education: str) -> None:
    """Функция, которая парсит расписание всех курсов и сохраняет их в json"""
    async with db_helper.session_factory() as session:
        schedule_path = settings.schedule.path.format(schedule_dir=form_education)
        file_path = Path(f"{BASE_DIR}/{schedule_path}")
        filenames = await aiofiles.os.listdir(file_path)
        service = ScheduleService(session=session)

        for filename in filenames:
            if filename == settings.schedule.file_name or filename.endswith(".json"):
                continue

            try:
                logger.info(f"Начало процесса парсинга данных из файла {filename}")

                full_file_path = file_path / filename
                wb = await asyncio.to_thread(
                    op.load_workbook, str(full_file_path), data_only=True
                )
                sheet = wb.active
                sheet_columns = sheet.max_column
                sheet_rows = sheet.max_row

                day_column = 0
                time_column = 0
                aud_list = []
                groups_columns = defaultdict(str)

                # Цикл, который пробегает по первой строке и ищет группы, колонку дней и колонку времени
                for col in range(1, sheet_columns + 1):
                    cell_val = sheet.cell(row=1, column=col).value

                    if cell_val:
                        if "время" in cell_val.lower() or "часы" in cell_val.lower():
                            time_column = col
                        elif cell_val.lower().startswith(("день", "дни")):
                            day_column = col
                        elif "ауд" in cell_val.lower():
                            aud_list.append(col)
                        elif cell_val[:2].isupper():
                            group = cell_val
                            groups_columns[col] = group

                # Кэш недельного расписания для каждой группы, чтобы не дергать базу при добавлении предметов
                schedules_cache = {}
                for col, group in groups_columns.items():
                    schedule = await service.get_schedule(
                        form_education=form_education,
                        faculty=filename[:-5],
                        group=group,
                    )

                    if not schedule:
                        schedule = await service.add_schedule(
                            form_education=form_education,
                            faculty=filename[:-5],
                            group=group,
                        )
                        logger.info(f"Создано новое расписание для группы {group}")
                    else:
                        await service.clear_schedule_content(schedule_id=schedule.id)
                        logger.info(f"Очищено содержимое расписания для группы {group}")

                    schedules_cache[group] = schedule

                # Начало парсинга данных о предметах
                current_day = ""
                past_time = ""
                current_time = ""
                daily_queue_numbers = defaultdict(
                    int
                )  # Счетчик для нумерации предметов на день для каждой группы
                daily_schedules_cache = {}

                # С 3 строки начинается расписание, выше находятся группы
                for row in range(3, sheet_rows):
                    for col in range(1, sheet_columns):
                        cell_val = get_merged_value(ws=sheet, r=row, c=col)

                        if (
                            cell_val
                            and (
                                len(
                                    cleaned_val := re.sub(r"\s+", " ", cell_val).strip()
                                )
                                > 2  # >2 чтобы не брать такие значения как "н." или "ч."
                            )
                            and ("декан" not in cell_val.lower())
                        ):
                            if col == day_column and cell_val.lower() != current_day:
                                current_day = cell_val.lower()
                            elif col == time_column and cell_val != current_time:
                                if past_time == "":
                                    past_time = cell_val

                                current_time = cell_val
                            elif pattern.match(cleaned_val):
                                current_time = cell_val
                            elif (
                                col in groups_columns.keys()
                                and cell_val
                                != get_merged_value(ws=sheet, r=row - 1, c=col)
                            ):
                                group_name = groups_columns[col]
                                schedule = schedules_cache[group_name]

                                # Создаем день если не создан
                                cache_key = f"{group_name}_{current_day}"
                                if cache_key not in daily_schedules_cache.keys():
                                    daily_schedule = await service.add_daily_schedule(
                                        name=current_day,
                                        schedule_id=schedule.id,
                                    )
                                    daily_schedules_cache[cache_key] = daily_schedule
                                    daily_queue_numbers[cache_key] = 1
                                else:
                                    daily_schedule = daily_schedules_cache[cache_key]

                                aud = "Не указано"
                                for aud_col in aud_list:
                                    if col < aud_col:
                                        aud = sheet.cell(row=row, column=aud_col).value
                                        break

                                subject_parity = None
                                if cleaned_val.lower().startswith("ч."):
                                    subject_parity = "четная"
                                elif cleaned_val.lower().startswith("н."):
                                    subject_parity = "нечетная"

                                await service.add_subject(
                                    name=cleaned_val,
                                    queue_number=daily_queue_numbers[cache_key],
                                    parity=subject_parity,
                                    time=current_time,
                                    audience=aud if aud else "Не указано",
                                    teacher=None,
                                    daily_schedule_id=daily_schedule.id,
                                )

                                if current_time != past_time:
                                    past_time = current_time
                                    daily_queue_numbers[cache_key] += 1

                # один коммит для всех предметов файла
                await service.commit()
                logger.info(f"Парсинг файла {filename} завершен успешно")

            except Exception as e:
                await service.session.rollback()
                logger.error(
                    f"Ошибка при парсинге файла {filename}: {e}", exc_info=True
                )
                continue


async def start_parser():
    tasks = [parser(form_education=key) for key in settings.zgy.urls.keys()]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    errors = [r for r in results if isinstance(r, Exception)]
    if errors:
        logger.error(f"Ошибки при парсинге: {len(errors)}/{len(tasks)}")
        for error in errors:
            logger.error(f"Ошибка: {error}", exc_info=error)
