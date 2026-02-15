import re
import asyncio
import logging
from pathlib import Path
from collections import defaultdict

import aiofiles
from aiofiles.os import listdir
import openpyxl as op
from openpyxl.cell.cell import MergedCell

from core import settings, BASE_DIR
from core.db import Subjects, WeeklySchedules, AllSchedules
from parser.schedule_crud import add_schedule, add_subject, add_day

logger = logging.getLogger(__name__)

pattern = re.compile(
    r"^([01]?\d|2[0-3]):[0-5]\d:[0-5]\d - ([01]?\d|2[0-3]):[0-5]\d:[0-5]\d$"
)


def get_merged_value(ws, r: int, c: int) -> str | None:
    """Выдает значение ячейки даже если она объединенная"""

    cell = ws.cell(row=r, column=c)

    # обычная ячейка
    if not isinstance(cell, MergedCell):
        return cell.value

    # ищем merge-диапазон, в который она входит
    for cr in ws.merged_cells.ranges:
        if cr.min_row <= r <= cr.max_row and cr.min_col <= c <= cr.max_col:
            return ws.cell(cr.min_row, cr.min_col).value

    return None


async def parser(form_education: str) -> None:
    """Функция, которая парсит расписание всех курсов и сохраняет их в json"""

    base_schedule = AllSchedules()
    schedule_path = settings.schedule.path.format(schedule_dir=form_education)
    file_path = Path(f"{BASE_DIR}/{schedule_path}")
    filenames = await aiofiles.os.listdir(file_path)

    for filename in filenames:
        if filename == settings.schedule.file_name or filename.endswith(".json"):
            continue

        logger.info(f"Начало процесса парсинга данных из файла {filename}")

        full_file_path = file_path / filename
        wb = await asyncio.to_thread(
            op.load_workbook, str(full_file_path), data_only=True
        )
        sheet = wb.active
        sheet_columns = sheet.max_column
        sheet_rows = sheet.max_row

        day_column = 0
        time_column = 0
        aud_list = []
        groups_columns = defaultdict(str)

        # Цикл, который пробегает по первой строке и ищет группы, колонку дней и колонку времени
        for col in range(1, sheet_columns + 1):
            cell_val = sheet.cell(row=1, column=col).value

            if cell_val:
                if "время" in cell_val.lower() or "часы" in cell_val.lower():
                    time_column = col
                elif cell_val.lower().startswith(("день", "дни")):
                    day_column = col
                elif "ауд" in cell_val.lower():
                    aud_list.append(col)
                elif cell_val[:2].isupper():
                    group = cell_val
                    new_weekly_schedule = WeeklySchedules(group_name=group)
                    await asyncio.to_thread(
                        add_schedule,
                        base_schedule=base_schedule,
                        schedule=new_weekly_schedule,
                        faculty=filename[:-5],
                    )

                    logger.info(
                        f"Пустое недельное расписание для факультета {filename[:-5]} инициализировано успешно"
                    )

                    groups_columns[col] = group

        # Начало парсинга данных о предметах
        current_day = ""
        current_time = ""
        added_day_in_groups = defaultdict(list)
        # С 3 строки начинается расписание, выше находятся группы
        for row in range(3, sheet_rows):
            for col in range(1, sheet_columns):
                cell_val = get_merged_value(ws=sheet, r=row, c=col)

                if (
                    cell_val
                    and (len(cleaned_val := re.sub(r"\s+", " ", cell_val).strip()) > 2)
                    and ("декан" not in cell_val.lower())
                ):  # >2 чтобы не брать такие значения как "н." или "ч."
                    if col == day_column and cell_val != current_day:
                        current_day = cell_val
                    elif col == time_column and cell_val != current_time:
                        current_time = cell_val

                    elif pattern.match(cleaned_val):
                        current_time = cell_val

                    elif col in groups_columns.keys() and cell_val != get_merged_value(
                        ws=sheet, r=row - 1, c=col
                    ):
                        group_name = groups_columns[col]
                        if current_day not in added_day_in_groups[group_name]:
                            await asyncio.to_thread(
                                add_day,
                                base_schedule=base_schedule,
                                day_name=current_day,
                                group_name=group_name,
                                faculty=filename[:-5],
                            )
                            logger.info(
                                f"Пустой день {current_day} добавлен в группу {filename[:-5]}"
                            )
                            added_day_in_groups[group_name].append(current_day)

                        aud = "Не указано"
                        for aud_col in aud_list:
                            if col < aud_col:
                                aud = sheet.cell(row=row, column=aud_col).value
                                break

                        await asyncio.to_thread(
                            add_subject,
                            base_schedule=base_schedule,
                            subject=Subjects(
                                subject_name=cleaned_val,
                                audience=aud if aud else "Не указано",
                                time=current_time,
                            ),
                            group_name=group_name,
                            day_name=current_day,
                            faculty=filename[:-5],
                        )

                        logger.info(
                            f"Добавлен предмет {cleaned_val} к расписанию группы {group_name} факультета {filename[:-5]}"
                        )

    output_file = file_path / settings.schedule.final_schedule
    async with aiofiles.open(output_file, "w", encoding="utf-8") as f:
        logger.info("Начало процесса записи расписания в json файл")
        await f.write(base_schedule.model_dump_json(indent=4, ensure_ascii=False))
        logger.info("Конец процесса записи расписания в json файл")


async def start_parser():
    tasks = [parser(form_education=key) for key in settings.zgy.urls.keys()]
    results = await asyncio.gather(*tasks, return_exceptions=True)

    errors = [r for r in results if isinstance(r, Exception)]
    if errors:
        logger.error(f"Ошибки при парсинге: {len(errors)}/{len(tasks)}")
        for error in errors:
            logger.error(f"Ошибка: {error}", exc_info=error)
