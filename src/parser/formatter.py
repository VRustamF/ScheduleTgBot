import asyncio
import logging
from pathlib import Path

import xlrd
from xlrd import Book

import openpyxl
from openpyxl import Workbook

from core import settings, BASE_DIR

logger = logging.getLogger(__name__)


def find_schedule_boundaries(sheet) -> tuple[int, int]:
    """Находит начальную и конечную строки расписания"""

    start_row = 0
    end_row = sheet.nrows

    # ищем слово "дни" для того, чтобы задать стартовую строку с которой будет начинаться расписание
    for c in range(sheet.ncols):
        for r in range(sheet.nrows):
            cell_val = sheet.cell_value(r, c)

            if isinstance(cell_val, str) and (
                "дни" in cell_val.lower() or "день" in cell_val.lower()
            ):
                start_row = r

        if start_row != 0:
            break

    if start_row == 0:
        logger.warning(f"В {sheet.name} слово 'дни' не найдено")
        start_row = 7

    # Ищем последнюю строку расписания
    saturday_row = None
    saturday_col = None
    for c in range(sheet.ncols):
        if saturday_row is not None:
            break

        for r in range(start_row, sheet.nrows):
            cell_val = sheet.cell_value(r, c)

            if isinstance(cell_val, str) and ("суббота" in cell_val.lower()):
                saturday_row = r
                saturday_col = c
                break

    if saturday_row is None:
        logger.warning(f"В {sheet.name} слово 'суббота' не найдено")
        return start_row, sheet.nrows

    # Ищем merged cell, содержащий субботу
    end_row = saturday_row + 1  # По умолчанию одна строка после субботы
    for r1, r2, c1, c2 in sheet.merged_cells:
        if r1 <= saturday_row < r2 and c1 <= saturday_col < c2:
            end_row = r2
            break

    logger.info(
        f"В {sheet.name} границы расписания: "
        f"строки {start_row}-{end_row-1} (всего {end_row - start_row} строк)"
    )

    return start_row, end_row


def copy_merge_cells_to_xlsx(s_name: str, book: Book) -> Workbook:
    """Функция, которая переносит объединенные ячейки из xls в xlsx"""

    sheet = book.sheet_by_name(s_name)

    # создаём xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    start_row, end_row = find_schedule_boundaries(sheet=sheet)

    # копируем значения
    for r in range(start_row, end_row):
        for c in range(sheet.ncols):
            ws.cell(row=r - start_row + 1, column=c + 1).value = sheet.cell_value(r, c)

    # переносим merged cells
    for r1, r2, c1, c2 in sheet.merged_cells:
        # полностью выше данных -> пропускаем
        if r2 <= start_row or r1 >= end_row:
            continue

        # подрезаем merge, если он начинается выше 8 строки
        new_r1 = max(r1, start_row)
        new_r2 = min(r2, end_row)

        ws.merge_cells(
            start_row=new_r1 - start_row + 1,
            end_row=new_r2 - start_row,
            start_column=c1 + 1,
            end_column=c2,
        )
    return wb


async def formatter(form_education: str) -> None:
    """
    Позволяет конвертировать xls в xlsx без потери объединенных ячеек.
    Создает для каждого листа отдельную таблицу
    """

    schedule_path = settings.schedule.path.format(schedule_dir=form_education)
    file_path = Path(f"{BASE_DIR}/{schedule_path}")

    # читаем xls
    xls_file_path = file_path / settings.schedule.file_name
    book = await asyncio.to_thread(
        xlrd.open_workbook,
        filename=str(xls_file_path),
        formatting_info=True,
    )
    sheet_names = book.sheet_names()

    for s_name in sheet_names:
        logger.info(f"Начало копирования объединенных ячеек для факультета {s_name}")

        workbook: Workbook = await asyncio.to_thread(
            copy_merge_cells_to_xlsx,
            s_name=s_name,
            book=book,
        )

        xlsx_file_path = file_path / f"{s_name}.xlsx"
        await asyncio.to_thread(workbook.save, str(xlsx_file_path))
        logger.info(f"Файл {s_name} сохранен")


async def start_formatter():
    tasks = [formatter(form_education=key) for key in settings.zgy.urls.keys()]
    await asyncio.gather(*tasks, return_exceptions=True)
