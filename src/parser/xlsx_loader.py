import xlrd
import openpyxl

from data import settings, BASE_DIR

FILES_PATH = f"{BASE_DIR}{settings.schedule.path}"

# читаем xls
book = xlrd.open_workbook(
    filename=f"{FILES_PATH}{settings.schedule.file_name}",
    formatting_info=True
)
sheet_names = book.sheet_names()

for s_name in sheet_names:
    sheet = book.sheet_by_name(s_name)

    # создаём xlsx
    wb = openpyxl.Workbook()
    ws = wb.active

    # копируем значения (не обязательно, но обычно нужно)
    for r in range(7, sheet.nrows):
        for c in range(sheet.ncols):
            ws.cell(row=r-7+1, column=c+1).value = sheet.cell_value(r, c)

    # переносим merged cells
    for r1, r2, c1, c2 in sheet.merged_cells:
        # полностью выше данных → пропускаем
        if r2 <= 7:
            continue

        # подрезаем merge, если он начинается выше 8 строки
        new_r1 = max(r1, 7)
        new_r2 = r2

        ws.merge_cells(
            start_row=new_r1 - 7 + 1,
            end_row=new_r2 - 7,
            start_column=c1 + 1,
            end_column=c2
        )

    wb.save(f"{FILES_PATH}{s_name}.xlsx")